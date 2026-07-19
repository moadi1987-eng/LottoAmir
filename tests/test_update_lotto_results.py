import os
import subprocess
import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook

from scripts.update_lotto_results import (
    DrawRecord,
    UpdateError,
    parse_csv,
    read_workbook,
    select_current_era,
    update_from_bytes,
    validate_records,
    write_workbook_atomic,
)


def official_csv_bytes(rows, encoding="cp1255"):
    header = (
        "\u05de\u05e1\u05e4\u05e8 \u05d4\u05d2\u05e8\u05dc\u05d4,"
        "\u05ea\u05d0\u05e8\u05d9\u05da,1,2,3,4,5,6,\u05d7\u05d6\u05e7,\u05e4\u05e8\u05e1,\u05d6\u05d5\u05db\u05d9\u05dd"
    )
    text = "\r\n".join([header, *[",".join(map(str, row)) for row in rows]])
    return text.encode(encoding)


def record(draw_number, draw_date, numbers, strong):
    return DrawRecord(draw_number, draw_date, tuple(numbers), strong)


class CsvParsingTests(unittest.TestCase):
    def setUp(self):
        self.rows = [
            (3944, "11/07/2026", 7, 14, 28, 30, 31, 37, 1, 0, 0),
            (3943, "07/07/2026", 5, 15, 16, 27, 32, 37, 2, 0, 1),
            (3942, "04/07/2026", 6, 7, 18, 19, 23, 34, 1, 2, 0),
        ]

    def test_parses_cp1255_official_rows_into_typed_records(self):
        records = parse_csv(official_csv_bytes(self.rows))

        self.assertEqual(len(records), 3)
        self.assertEqual(
            records[0],
            record(3944, "11/07/2026", (7, 14, 28, 30, 31, 37), 1),
        )
        self.assertEqual(records[-1].draw_number, 3942)

    def test_parses_utf8_with_bom(self):
        records = parse_csv(official_csv_bytes(self.rows, encoding="utf-8-sig"))

        self.assertEqual(records[1].draw_date, "07/07/2026")

    def test_rejects_malformed_data_row(self):
        bad_rows = [(3944, "11/07/2026", 7, 14, 28, "bad", 31, 37, 1, 0, 0)]

        with self.assertRaisesRegex(UpdateError, "CSV row 2"):
            parse_csv(official_csv_bytes(bad_rows))


class EraSelectionTests(unittest.TestCase):
    def setUp(self):
        self.records = [
            record(3944, "11/07/2026", (7, 14, 28, 30, 31, 37), 1),
            record(3943, "07/07/2026", (5, 15, 16, 27, 32, 37), 2),
            record(3942, "04/07/2026", (6, 7, 18, 19, 23, 34), 1),
            record(9934, "30/06/2026", (1, 2, 3, 4, 5, 49), 49),
        ]

    def test_selects_only_contiguous_prefix_down_to_existing_oldest_draw(self):
        selected = select_current_era(self.records, oldest_draw=3942)

        self.assertEqual([item.draw_number for item in selected], [3944, 3943, 3942])

    def test_rejects_gap_before_existing_oldest_draw(self):
        records_with_gap = [self.records[0], self.records[2]]

        with self.assertRaisesRegex(UpdateError, "expected draw 3943"):
            select_current_era(records_with_gap, oldest_draw=3942)

    def test_rejects_missing_existing_oldest_draw(self):
        with self.assertRaisesRegex(UpdateError, "did not contain existing oldest draw 3941"):
            select_current_era(self.records[:3], oldest_draw=3941)


class RecordValidationTests(unittest.TestCase):
    def valid_records(self):
        return [
            record(3944, "11/07/2026", (7, 14, 28, 30, 31, 37), 1),
            record(3943, "07/07/2026", (5, 15, 16, 27, 32, 37), 8),
        ]

    def test_accepts_historical_strong_eight(self):
        validate_records(self.valid_records(), existing_newest=3943, existing_count=1)

    def test_rejects_duplicate_regular_numbers(self):
        rows = self.valid_records()
        rows[1] = record(3943, "07/07/2026", (5, 5, 16, 27, 32, 37), 2)

        with self.assertRaisesRegex(UpdateError, "six unique regular numbers"):
            validate_records(rows, existing_newest=3943, existing_count=1)

    def test_rejects_regular_number_outside_current_game_range(self):
        rows = self.valid_records()
        rows[1] = record(3943, "07/07/2026", (5, 15, 16, 27, 32, 38), 2)

        with self.assertRaisesRegex(UpdateError, "regular number outside 1..37"):
            validate_records(rows, existing_newest=3943, existing_count=1)

    def test_rejects_historical_strong_outside_one_through_eight(self):
        rows = self.valid_records()
        rows[1] = record(3943, "07/07/2026", (5, 15, 16, 27, 32, 37), 9)

        with self.assertRaisesRegex(UpdateError, "strong number outside 1..8"):
            validate_records(rows, existing_newest=3943, existing_count=1)

    def test_rejects_newest_strong_eight(self):
        rows = self.valid_records()
        rows[0] = record(3944, "11/07/2026", (7, 14, 28, 30, 31, 37), 8)

        with self.assertRaisesRegex(UpdateError, "newest draw strong number outside 1..7"):
            validate_records(rows, existing_newest=3943, existing_count=1)

    def test_rejects_download_older_or_shorter_than_workbook(self):
        with self.assertRaisesRegex(UpdateError, "older than workbook"):
            validate_records(self.valid_records(), existing_newest=3945, existing_count=1)

        with self.assertRaisesRegex(UpdateError, "shorter than workbook"):
            validate_records(self.valid_records(), existing_newest=3943, existing_count=3)


class WorkbookUpdateTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.workbook_path = Path(self.temp_dir.name) / "NUMBERS.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "111"
        sheet.append([3943, datetime(2026, 7, 7), 5, 15, 16, 27, 32, 37, 2])
        sheet.append([3942, "04/07/2026", 6, 7, 18, 19, 23, 34, 1])
        sheet["A10"].number_format = "0"
        workbook.save(self.workbook_path)

        self.csv_bytes = official_csv_bytes(
            [
                (3944, "11/07/2026", 7, 14, 28, 30, 31, 37, 1, 0, 0),
                (3943, "07/07/2026", 5, 15, 16, 27, 32, 37, 2, 0, 1),
                (3942, "04/07/2026", 6, 7, 18, 19, 23, 34, 1, 2, 0),
            ]
        )

    def test_updates_workbook_atomically_in_existing_format(self):
        changed = update_from_bytes(self.workbook_path, self.csv_bytes)

        self.assertTrue(changed)
        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        self.assertEqual(workbook.sheetnames, ["111"])
        sheet = workbook["111"]
        values = list(sheet.iter_rows(values_only=True))
        self.assertEqual(len(values), 3)
        self.assertEqual(values[0], (3944, "11/07/2026", 7, 14, 28, 30, 31, 37, 1))
        self.assertEqual(values[-1][0], 3942)
        self.assertTrue(all(len(row) == 9 for row in values))
        workbook.close()

        styled_workbook = load_workbook(self.workbook_path, read_only=False)
        styled_sheet = styled_workbook["111"]
        self.assertEqual(styled_sheet.column_dimensions["A"].width, 10)
        self.assertEqual(styled_sheet.column_dimensions["B"].width, 14)
        for column in "CDEFGHI":
            self.assertEqual(styled_sheet.column_dimensions[column].width, 7)
        styled_workbook.close()

    def test_does_not_rewrite_identical_normalized_data(self):
        self.assertTrue(update_from_bytes(self.workbook_path, self.csv_bytes))
        first_bytes = self.workbook_path.read_bytes()

        changed = update_from_bytes(self.workbook_path, self.csv_bytes)

        self.assertFalse(changed)
        self.assertEqual(self.workbook_path.read_bytes(), first_bytes)

    def test_invalid_download_leaves_existing_workbook_untouched(self):
        original_bytes = self.workbook_path.read_bytes()
        bad_csv = official_csv_bytes(
            [(3944, "11/07/2026", 7, 7, 28, 30, 31, 37, 1, 0, 0)]
        )

        with self.assertRaises(UpdateError):
            update_from_bytes(self.workbook_path, bad_csv)

        self.assertEqual(self.workbook_path.read_bytes(), original_bytes)

    def test_rejects_valid_looking_changes_to_existing_draw_history(self):
        original_bytes = self.workbook_path.read_bytes()
        changed_history = official_csv_bytes(
            [
                (3944, "11/07/2026", 7, 14, 28, 30, 31, 37, 1, 0, 0),
                (3943, "07/07/2026", 4, 15, 16, 27, 32, 37, 2, 0, 1),
                (3942, "04/07/2026", 6, 7, 18, 19, 23, 34, 1, 2, 0),
            ]
        )

        with self.assertRaisesRegex(UpdateError, "changed existing draw 3943"):
            update_from_bytes(self.workbook_path, changed_history)

        self.assertEqual(self.workbook_path.read_bytes(), original_bytes)

    def test_replace_failure_leaves_original_workbook_and_cleans_temporary_file(self):
        original_bytes = self.workbook_path.read_bytes()
        snapshot = read_workbook(self.workbook_path)

        with mock.patch(
            "scripts.update_lotto_results.os.replace",
            side_effect=OSError("simulated replace failure"),
        ):
            with self.assertRaisesRegex(UpdateError, "Could not write workbook atomically"):
                write_workbook_atomic(
                    self.workbook_path, snapshot.records, snapshot.sheet_name
                )

        self.assertEqual(self.workbook_path.read_bytes(), original_bytes)
        self.assertEqual(list(self.workbook_path.parent.glob(".NUMBERS-*.xlsx")), [])

    def test_temporary_save_failure_leaves_original_workbook_untouched(self):
        original_bytes = self.workbook_path.read_bytes()
        snapshot = read_workbook(self.workbook_path)

        with mock.patch.object(
            Workbook, "save", side_effect=OSError("simulated save failure")
        ):
            with self.assertRaisesRegex(UpdateError, "Could not write workbook atomically"):
                write_workbook_atomic(
                    self.workbook_path, snapshot.records, snapshot.sheet_name
                )

        self.assertEqual(self.workbook_path.read_bytes(), original_bytes)
        self.assertEqual(list(self.workbook_path.parent.glob(".NUMBERS-*.xlsx")), [])

    def test_temporary_verification_failure_leaves_original_workbook_untouched(self):
        original_bytes = self.workbook_path.read_bytes()
        snapshot = read_workbook(self.workbook_path)
        mismatched_snapshot = mock.Mock(sheet_name="111", records=())

        with mock.patch(
            "scripts.update_lotto_results.read_workbook",
            return_value=mismatched_snapshot,
        ):
            with self.assertRaisesRegex(
                UpdateError, "Temporary workbook verification did not match"
            ):
                write_workbook_atomic(
                    self.workbook_path, snapshot.records, snapshot.sheet_name
                )

        self.assertEqual(self.workbook_path.read_bytes(), original_bytes)
        self.assertEqual(list(self.workbook_path.parent.glob(".NUMBERS-*.xlsx")), [])


class RepositoryWorkbookContractTests(unittest.TestCase):
    def test_repository_workbook_keeps_the_current_game_era_boundary(self):
        workbook_path = Path(__file__).resolve().parents[1] / "NUMBERS.xlsx"
        snapshot = read_workbook(workbook_path)

        self.assertEqual(snapshot.sheet_name, "111")
        self.assertGreaterEqual(snapshot.newest_draw, 3944)
        self.assertEqual(snapshot.oldest_draw, 2233)
        self.assertEqual(
            snapshot.record_count, snapshot.newest_draw - snapshot.oldest_draw + 1
        )

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        sheet = workbook[snapshot.sheet_name]
        self.assertEqual(sheet.max_column, 9)
        workbook.close()


class LocalSchedulerContractTests(unittest.TestCase):
    def test_windows_task_uses_isolated_clone_and_pushes_only_new_results(self):
        root = Path(__file__).resolve().parents[1]
        runner_path = root / "scripts" / "run_scheduled_update.ps1"
        installer_path = root / "scripts" / "install_lotto_update_task.ps1"
        requirements_path = root / "scripts" / "requirements-lotto-update.txt"

        self.assertTrue(runner_path.is_file(), "Scheduled runner must exist")
        self.assertTrue(installer_path.is_file(), "Task installer must exist")
        self.assertTrue(requirements_path.is_file(), "Updater requirements must exist")

        runner = runner_path.read_text(encoding="utf-8")
        runner_fragments = (
            "LottoAmirUpdater",
            "Archive-AutomationClone",
            "git clone",
            "git pull --ff-only origin main",
            "origin/main...main",
            '$AllowedDataPaths = @("LOTTO_PRIZES.json", "NUMBERS.xlsx")',
            "Test-OnlyAllowedDataPaths",
            "scripts/update_lotto_results.py",
            "scripts/update_lotto_prizes.py",
            "git add -- $AllowedDataPaths",
            "git push origin main",
        )
        for fragment in runner_fragments:
            with self.subTest(fragment=fragment):
                self.assertIn(fragment, runner)

        installer = installer_path.read_text(encoding="utf-8")
        installer_fragments = (
            "New-ScheduledTaskTrigger",
            "-Weekly",
            "-DaysOfWeek Tuesday, Thursday, Saturday",
            '-At "23:55"',
            "New-ScheduledTaskSettingsSet",
            "-StartWhenAvailable",
            "Register-ScheduledTask",
            "run_scheduled_update.ps1",
            "$env:SystemRoot",
        )
        for fragment in installer_fragments:
            with self.subTest(fragment=fragment):
                self.assertIn(fragment, installer)
        self.assertNotIn("New-TimeSpan -Hours 6", installer)
        self.assertNotIn("New-ScheduledTaskTrigger -AtLogOn", installer)

    def test_blocked_github_hosted_updater_workflows_are_removed(self):
        workflows = Path(__file__).resolve().parents[1] / ".github" / "workflows"

        self.assertFalse((workflows / "update-lotto-results.yml").exists())
        self.assertFalse((workflows / "diagnose-pais.yml").exists())


@unittest.skipUnless(os.name == "nt", "Windows Task Scheduler runner test")
class WindowsSchedulerRecoveryTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.root = Path(self.temp_dir.name)
        self.remote = self.root / "remote.git"
        self.seed = self.root / "seed"
        self.automation_root = self.root / "automation"
        self.runner = (
            Path(__file__).resolve().parents[1]
            / "scripts"
            / "run_scheduled_update.ps1"
        )
        self.power_shell = (
            Path(os.environ["SystemRoot"])
            / "System32"
            / "WindowsPowerShell"
            / "v1.0"
            / "powershell.exe"
        )

        self.run_command(
            ["git", "init", "--bare", "--initial-branch=main", str(self.remote)]
        )
        self.run_command(
            ["git", "init", "--initial-branch=main", str(self.seed)]
        )
        self.run_command(["git", "config", "user.name", "Test User"], self.seed)
        self.run_command(
            ["git", "config", "user.email", "test@example.com"], self.seed
        )
        (self.seed / "scripts").mkdir()
        (self.seed / "scripts" / "update_lotto_results.py").write_text(
            "print('fixture updater: no changes')\n", encoding="utf-8"
        )
        (self.seed / "scripts" / "update_lotto_prizes.py").write_text(
            "print('fixture prize updater: no changes')\n", encoding="utf-8"
        )
        (self.seed / "NUMBERS.xlsx").write_text("3944\n", encoding="utf-8")
        (self.seed / "LOTTO_PRIZES.json").write_text(
            '{"draws":{},"schemaVersion":1,"updatedAt":null}\n', encoding="utf-8"
        )
        (self.seed / "README.md").write_text("initial\n", encoding="utf-8")
        self.run_command(["git", "add", "."], self.seed)
        self.run_command(["git", "commit", "-m", "initial"], self.seed)
        self.run_command(
            ["git", "remote", "add", "origin", str(self.remote)], self.seed
        )
        self.run_command(["git", "push", "-u", "origin", "main"], self.seed)

    def run_command(self, command, cwd=None):
        return subprocess.run(
            command,
            cwd=cwd,
            check=True,
            text=True,
            capture_output=True,
        )

    def run_scheduler(self):
        return subprocess.run(
            [
                str(self.power_shell),
                "-NoProfile",
                "-NonInteractive",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(self.runner),
                "-AutomationRoot",
                str(self.automation_root),
                "-RepositoryUrl",
                str(self.remote),
                "-PythonExecutable",
                sys.executable,
                "-NoPush",
            ],
            check=False,
            text=True,
            capture_output=True,
        )

    def test_recovers_allowed_data_changes_and_diverged_data_commit(self):
        first_run = self.run_scheduler()
        self.assertEqual(first_run.returncode, 0, first_run.stdout + first_run.stderr)

        managed_repo = self.automation_root / "repo"
        (managed_repo / "NUMBERS.xlsx").write_text("interrupted\n", encoding="utf-8")
        (managed_repo / "LOTTO_PRIZES.json").write_text(
            "interrupted\n", encoding="utf-8"
        )

        dirty_recovery = self.run_scheduler()
        self.assertEqual(
            dirty_recovery.returncode,
            0,
            dirty_recovery.stdout + dirty_recovery.stderr,
        )
        self.assertEqual(
            (managed_repo / "NUMBERS.xlsx").read_text(encoding="utf-8"), "3944\n"
        )
        self.assertIn(
            '"schemaVersion":1',
            (managed_repo / "LOTTO_PRIZES.json").read_text(encoding="utf-8"),
        )

        self.run_command(["git", "config", "user.name", "Updater"], managed_repo)
        self.run_command(
            ["git", "config", "user.email", "updater@example.com"], managed_repo
        )
        (managed_repo / "NUMBERS.xlsx").write_text("pending push\n", encoding="utf-8")
        (managed_repo / "LOTTO_PRIZES.json").write_text(
            "pending prizes\n", encoding="utf-8"
        )
        self.run_command(
            ["git", "add", "NUMBERS.xlsx", "LOTTO_PRIZES.json"], managed_repo
        )
        self.run_command(["git", "commit", "-m", "pending data"], managed_repo)

        remote_writer = self.root / "remote-writer"
        self.run_command(["git", "clone", str(self.remote), str(remote_writer)])
        self.run_command(["git", "config", "user.name", "Remote User"], remote_writer)
        self.run_command(
            ["git", "config", "user.email", "remote@example.com"], remote_writer
        )
        (remote_writer / "README.md").write_text("remote advance\n", encoding="utf-8")
        self.run_command(["git", "add", "README.md"], remote_writer)
        self.run_command(["git", "commit", "-m", "remote advance"], remote_writer)
        self.run_command(["git", "push", "origin", "main"], remote_writer)

        divergence_recovery = self.run_scheduler()
        self.assertEqual(
            divergence_recovery.returncode,
            0,
            divergence_recovery.stdout + divergence_recovery.stderr,
        )
        self.assertEqual(
            (managed_repo / "README.md").read_text(encoding="utf-8"),
            "remote advance\n",
        )
        self.assertGreaterEqual(
            len(list(self.automation_root.glob("repo-recovery-*"))), 2
        )


if __name__ == "__main__":
    unittest.main()
