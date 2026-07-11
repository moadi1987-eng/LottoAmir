#!/usr/bin/env python3
"""Safely refresh NUMBERS.xlsx from the official Pais lottery CSV."""

from __future__ import annotations

import argparse
import csv
import io
import os
import sys
import tempfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Sequence
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook


DEFAULT_CSV_URL = "https://www.pais.co.il/Lotto/lotto_resultsDownload.aspx"
ARCHIVE_URL = "https://www.pais.co.il/lotto/archive.aspx"
DEFAULT_WORKBOOK = Path("NUMBERS.xlsx")
DOWNLOAD_TIMEOUT_SECONDS = 30


class UpdateError(RuntimeError):
    """Raised when source data or workbook state is unsafe to publish."""


@dataclass(frozen=True)
class DrawRecord:
    draw_number: int
    draw_date: str
    regular_numbers: tuple[int, int, int, int, int, int]
    strong_number: int

    def as_row(self) -> list[int | str]:
        return [
            self.draw_number,
            self.draw_date,
            *self.regular_numbers,
            self.strong_number,
        ]


@dataclass(frozen=True)
class WorkbookSnapshot:
    sheet_name: str
    records: tuple[DrawRecord, ...]

    @property
    def newest_draw(self) -> int:
        return self.records[0].draw_number

    @property
    def oldest_draw(self) -> int:
        return self.records[-1].draw_number

    @property
    def record_count(self) -> int:
        return len(self.records)


def decode_csv(payload: bytes) -> str:
    if not payload:
        raise UpdateError("Official CSV download was empty")

    for encoding in ("utf-8-sig", "cp1255", "iso-8859-8"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise UpdateError("Official CSV could not be decoded")


def normalize_draw_date(value: object, context: str) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")

    text = str(value).strip()
    try:
        parsed = datetime.strptime(text, "%d/%m/%Y")
    except (TypeError, ValueError) as exc:
        raise UpdateError(f"{context}: invalid date {text!r}") from exc
    return parsed.strftime("%d/%m/%Y")


def coerce_integer(value: object, context: str) -> int:
    if isinstance(value, bool):
        raise UpdateError(f"{context}: expected an integer, got {value!r}")
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)

    text = str(value).strip()
    try:
        return int(text)
    except (TypeError, ValueError) as exc:
        raise UpdateError(f"{context}: expected an integer, got {value!r}") from exc


def parse_csv(payload: bytes) -> list[DrawRecord]:
    text = decode_csv(payload)
    records: list[DrawRecord] = []

    for row_number, row in enumerate(csv.reader(io.StringIO(text)), start=1):
        if not row or not any(cell.strip() for cell in row):
            continue

        first_cell = row[0].strip().lstrip("\ufeff")
        if not first_cell.isdecimal():
            if records:
                raise UpdateError(f"CSV row {row_number}: unexpected non-data row")
            continue
        if len(row) < 9:
            raise UpdateError(f"CSV row {row_number}: expected at least 9 columns")

        try:
            draw_number = coerce_integer(first_cell, f"CSV row {row_number} draw")
            draw_date = normalize_draw_date(row[1], f"CSV row {row_number}")
            regular = tuple(
                coerce_integer(value, f"CSV row {row_number} regular number")
                for value in row[2:8]
            )
            strong = coerce_integer(row[8], f"CSV row {row_number} strong number")
        except UpdateError as exc:
            raise UpdateError(f"CSV row {row_number}: {exc}") from exc

        records.append(
            DrawRecord(
                draw_number=draw_number,
                draw_date=draw_date,
                regular_numbers=regular,  # type: ignore[arg-type]
                strong_number=strong,
            )
        )

    if not records:
        raise UpdateError("Official CSV contained no draw rows")
    return records


def select_current_era(
    records: Sequence[DrawRecord], oldest_draw: int
) -> list[DrawRecord]:
    if not records:
        raise UpdateError("Official CSV contained no draw rows")
    if oldest_draw <= 0:
        raise UpdateError("Workbook oldest draw must be positive")

    selected: list[DrawRecord] = []
    expected_draw = records[0].draw_number
    for item in records:
        if item.draw_number != expected_draw:
            raise UpdateError(
                "Official CSV sequence broke before the workbook boundary: "
                f"expected draw {expected_draw}, got {item.draw_number}"
            )
        selected.append(item)
        if item.draw_number == oldest_draw:
            return selected
        expected_draw -= 1

    raise UpdateError(
        f"Official CSV did not contain existing oldest draw {oldest_draw}"
    )


def validate_records(
    records: Sequence[DrawRecord], existing_newest: int, existing_count: int
) -> None:
    if not records:
        raise UpdateError("Selected official dataset was empty")
    if records[0].draw_number < existing_newest:
        raise UpdateError(
            f"Official newest draw {records[0].draw_number} is older than workbook "
            f"draw {existing_newest}"
        )
    if len(records) < existing_count:
        raise UpdateError(
            f"Official selected range is shorter than workbook: {len(records)} < "
            f"{existing_count}"
        )

    draw_numbers = [item.draw_number for item in records]
    if len(draw_numbers) != len(set(draw_numbers)):
        raise UpdateError("Selected official dataset contains duplicate draw numbers")
    for current, following in zip(records, records[1:]):
        if current.draw_number - following.draw_number != 1:
            raise UpdateError(
                "Selected official dataset is not contiguous: "
                f"{current.draw_number} then {following.draw_number}"
            )

    for item in records:
        normalize_draw_date(item.draw_date, f"draw {item.draw_number}")
        regular = item.regular_numbers
        if len(regular) != 6 or len(set(regular)) != 6:
            raise UpdateError(
                f"Draw {item.draw_number} must contain six unique regular numbers"
            )
        if any(type(number) is not int or number < 1 or number > 37 for number in regular):
            raise UpdateError(
                f"Draw {item.draw_number} has a regular number outside 1..37"
            )
        if type(item.strong_number) is not int or not 1 <= item.strong_number <= 8:
            raise UpdateError(
                f"Draw {item.draw_number} has a strong number outside 1..8"
            )

    if not 1 <= records[0].strong_number <= 7:
        raise UpdateError(
            f"Draw {records[0].draw_number} has newest draw strong number outside 1..7"
        )


def validate_existing_history(
    official_records: Sequence[DrawRecord], existing_records: Sequence[DrawRecord]
) -> None:
    official_by_draw = {item.draw_number: item for item in official_records}
    for existing in existing_records:
        official = official_by_draw.get(existing.draw_number)
        if official is None:
            raise UpdateError(
                f"Official source omitted existing draw {existing.draw_number}"
            )
        if official != existing:
            raise UpdateError(
                f"Official source changed existing draw {existing.draw_number}; "
                "automatic history rewrites are blocked"
            )


def read_workbook(path: Path | str) -> WorkbookSnapshot:
    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise UpdateError(f"Workbook does not exist: {workbook_path}")

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        raise UpdateError(f"Could not open workbook {workbook_path}: {exc}") from exc

    try:
        if len(workbook.sheetnames) != 1:
            raise UpdateError(
                f"Workbook must contain exactly one sheet, found {len(workbook.sheetnames)}"
            )
        sheet_name = workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        records: list[DrawRecord] = []

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_col=9, values_only=True), start=1
        ):
            if not any(value is not None and value != "" for value in row):
                continue
            if row[0] is None or row[0] == "":
                raise UpdateError(
                    f"Workbook row {row_number} has data without a draw number"
                )

            draw_number = coerce_integer(row[0], f"Workbook row {row_number} draw")
            draw_date = normalize_draw_date(row[1], f"Workbook row {row_number}")
            regular = tuple(
                coerce_integer(value, f"Workbook row {row_number} regular number")
                for value in row[2:8]
            )
            strong = coerce_integer(
                row[8], f"Workbook row {row_number} strong number"
            )
            records.append(
                DrawRecord(
                    draw_number=draw_number,
                    draw_date=draw_date,
                    regular_numbers=regular,  # type: ignore[arg-type]
                    strong_number=strong,
                )
            )
    finally:
        workbook.close()

    if not records:
        raise UpdateError("Workbook contains no draw rows")
    validate_records(records, records[0].draw_number, len(records))
    return WorkbookSnapshot(sheet_name=sheet_name, records=tuple(records))


def write_workbook_atomic(
    path: Path | str, records: Iterable[DrawRecord], sheet_name: str
) -> None:
    workbook_path = Path(path)
    rows = tuple(records)
    if not rows:
        raise UpdateError("Refusing to write an empty workbook")

    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{workbook_path.stem}-",
            suffix=".xlsx",
            dir=workbook_path.parent,
            delete=False,
        ) as handle:
            temp_path = Path(handle.name)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.column_dimensions["A"].width = 10
        sheet.column_dimensions["B"].width = 14
        for column in "CDEFGHI":
            sheet.column_dimensions[column].width = 7
        for item in rows:
            sheet.append(item.as_row())
        workbook.save(temp_path)
        workbook.close()

        verification = read_workbook(temp_path)
        if verification.sheet_name != sheet_name or verification.records != rows:
            raise UpdateError("Temporary workbook verification did not match source data")

        os.replace(temp_path, workbook_path)
        temp_path = None
    except UpdateError:
        raise
    except Exception as exc:
        raise UpdateError(f"Could not write workbook atomically: {exc}") from exc
    finally:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)


def update_from_bytes(path: Path | str, payload: bytes) -> bool:
    workbook_path = Path(path)
    existing = read_workbook(workbook_path)
    all_official_records = parse_csv(payload)
    selected = select_current_era(all_official_records, existing.oldest_draw)
    validate_records(selected, existing.newest_draw, existing.record_count)
    validate_existing_history(selected, existing.records)

    if tuple(selected) == existing.records:
        return False

    write_workbook_atomic(workbook_path, selected, existing.sheet_name)
    return True


def download_csv(
    url: str = DEFAULT_CSV_URL, timeout: int = DOWNLOAD_TIMEOUT_SECONDS
) -> bytes:
    request = Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (compatible; LottoAmir-Updater/1.0)",
            "Referer": ARCHIVE_URL,
            "Accept": "text/csv,text/plain,application/octet-stream,*/*",
        },
    )
    try:
        with urlopen(request, timeout=timeout) as response:
            status = getattr(response, "status", 200)
            if status != 200:
                raise UpdateError(f"Official CSV returned HTTP {status}")
            payload = response.read()
    except HTTPError as exc:
        raise UpdateError(f"Official CSV returned HTTP {exc.code}") from exc
    except URLError as exc:
        raise UpdateError(f"Official CSV download failed: {exc.reason}") from exc
    except TimeoutError as exc:
        raise UpdateError("Official CSV download timed out") from exc

    if not payload:
        raise UpdateError("Official CSV download was empty")
    return payload


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Safely update NUMBERS.xlsx from the official Pais CSV"
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help="Workbook to update (default: NUMBERS.xlsx)",
    )
    parser.add_argument(
        "--csv-file",
        type=Path,
        help="Use a local CSV instead of downloading the official endpoint",
    )
    parser.add_argument(
        "--url",
        default=DEFAULT_CSV_URL,
        help="CSV URL used when --csv-file is not supplied",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_argument_parser().parse_args(argv)
    try:
        payload = args.csv_file.read_bytes() if args.csv_file else download_csv(args.url)
        changed = update_from_bytes(args.workbook, payload)
        snapshot = read_workbook(args.workbook)
        state = "updated" if changed else "already current"
        print(
            f"NUMBERS.xlsx {state}: newest draw {snapshot.newest_draw}, "
            f"{snapshot.record_count} rows"
        )
        return 0
    except (OSError, UpdateError) as exc:
        print(f"Lotto results update failed: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
